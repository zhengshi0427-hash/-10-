import requests
import pandas as pd
from lxml import etree
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ==================== 配置 h=好评 ====================
base_url = "https://movie.douban.com/subject/37116446/comments?percent_type=h&limit=20&status=P&sort=new_score&start="

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Referer": "https://movie.douban.com/subject/37116446/",
    "Cookie": 'll="118172"; bid=IhGRS2uQwII; dbcl2="295327634:oTNuHdaewsA";'
}

csv_columns = ["用户链接", "用户名", "点赞数", "星级评分", "评论时间", "发布地区", "评论内容"]
all_data = []

star_map = {
    "力荐": "5星",
    "推荐": "4星",
    "还行": "3星",
    "较差": "2星",
    "很差": "1星",
    "未评分": ""
}

# ==================== 解析函数（适配页面结构） ====================
def parse_page(html):
    tree = etree.HTML(html)
    items = tree.xpath('//div[@class="comment-item"]//div[@class="comment"]')
    if len(items) == 0:
        print("⚠️ 当前页面无评论，停止翻页")
        return []
    print(f"本页共{len(items)}条好评")
    for idx, item in enumerate(items,1):
        time.sleep(random.uniform(0.2,0.5))
        data = {}
        # 用户链接
        data["用户链接"] = item.xpath('.//span[@class="comment-info"]/a/@href')[0] if item.xpath('.//span[@class="comment-info"]/a/@href') else ""
        # 用户名
        data["用户名"] = item.xpath('.//span[@class="comment-info"]/a/text()')[0].strip() if item.xpath('.//span[@class="comment-info"]/a/text()') else ""
        # 点赞数 vote-count
        data["点赞数"] = item.xpath('.//span[contains(@class,"vote-count")]/text()')[0].strip() if item.xpath('.//span[contains(@class,"vote-count")]/text()') else "0"
        # 星级
        raw_star = item.xpath('.//span[contains(@class,"allstar")]/@title')
        data["星级评分"] = star_map.get(raw_star[0] if raw_star else "未评分","")
        # 评论时间（取a标签title）
        data["评论时间"] = item.xpath('.//a[@class="comment-time"]/@title')[0].strip() if item.xpath('.//a[@class="comment-time"]/@title') else ""
        # 发布地区
        data["发布地区"] = item.xpath('.//span[@class="comment-location"]/text()')[0].strip() if item.xpath('.//span[@class="comment-location"]/text()') else ""
        # 短评正文
        data["评论内容"] = item.xpath('string(.//span[@class="short"])').strip()

        all_data.append(data)
        print(f"{idx} | {data['用户名']} | {data['星级评分']} | {data['评论时间'][:10]} | {data['评论内容'][:22]}...")
    return items

# ==================== 自动翻页主程序 ====================
if __name__ == "__main__":
    start = 0
    try:
        while True:
            url = base_url + str(start)
            print(f"\n===== 正在抓取start={start} 好评页面 =====")
            resp = requests.get(url,headers=headers,timeout=12)
            res = parse_page(resp.text)
            if not res:
                break
            start += 20
            time.sleep(random.uniform(1.5,2.5))
    except KeyboardInterrupt:
        print("\n⚠️ 手动终止，保存已抓取数据")
    finally:
        # 导出CSV
        df = pd.DataFrame(all_data,columns=csv_columns)
        df.to_csv("豆瓣_好评_h.csv",index=False,encoding="utf-8-sig")
        # 生成自动换行Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "好评短评"
        ws.append(csv_columns)
        for d in all_data:
            ws.append(list(d.values()))
        col_width_cfg = {"A":30,"B":14,"C":10,"D":10,"E":20,"F":10,"G":70}
        for col,w in col_width_cfg.items():
            ws.column_dimensions[col].width = w
        wrap_align = Alignment(wrap_text=True,vertical="top")
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=7):
            for cell in row:
                cell.alignment = wrap_align
        wb.save("豆瓣_好评_h.xlsx")
        print(f"\n✅ 抓取完成，合计{len(all_data)}条，文件已保存")