import requests
import pandas as pd
from lxml import etree
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ==================== 配置 ====================
base_url = "https://movie.douban.com/subject/37116446/reviews?rating="
target_count = 300  # 爬300条

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Referer": "https://movie.douban.com/subject/37116446/",
    "Cookie": 'll="118172"; bid=IhGRS2uQwII; dbcl2="295327634:oTNuHdaewsA";'
}

# 6 个字段（已删除地区）
csv_columns = ["用户链接", "用户名", "点赞数", "星级评分", "评论时间", "评论内容"]
all_data = []

star_map = {
    "力荐": "5星",
    "推荐": "4星",
    "还行": "3星",
    "较差": "2星",
    "很差": "1星",
    "未评分": ""
}

# 获取完整影评
def get_full_review(rid):
    try:
        api_url = f"https://movie.douban.com/j/review/{rid}/full"
        resp = requests.get(api_url, headers=headers, timeout=5)
        if resp.status_code == 200:
            html = resp.json()["html"]
            tree = etree.HTML(html)
            return tree.xpath("string()").strip()
    except:
        return ""
    return ""

# ==================== 解析 ====================
def parse_page(html):
    tree = etree.HTML(html)
    items = tree.xpath('//div[contains(@class,"review-item")]')

    for item in items:
        if len(all_data) >= target_count:
            break

        time.sleep(random.uniform(0.5, 1))
        data = {}

        # 1 用户链接
        data["用户链接"] = item.xpath('.//a[@class="avator"]/@href')[0] if item.xpath('.//a[@class="avator"]/@href') else ""

        # 2 用户名
        data["用户名"] = item.xpath('.//a[@class="name"]/text()')[0].strip() if item.xpath('.//a[@class="name"]/text()') else ""

        # 3 点赞数
        data["点赞数"] = item.xpath('.//a[contains(@class,"up")]/text()')[0].strip() if item.xpath('.//a[contains(@class,"up")]/text()') else "0"

        # 4 星级评分
        raw_star = item.xpath('.//span[contains(@class,"allstar")]/@title')
        data["星级评分"] = star_map.get(raw_star[0] if raw_star else "未评分", "")

        # 5 评论时间
        data["评论时间"] = item.xpath('.//span[@class="main-meta"]/text()')[0].strip() if item.xpath('.//span[@class="main-meta"]/text()') else ""

        # 6 评论内容（自动展开全文）
        rid = item.xpath('.//div[@class="review-short"]/@data-rid')
        if rid:
            data["评论内容"] = get_full_review(rid[0])
        else:
            data["评论内容"] = item.xpath('string(.//div[@class="short-content"])').strip()

        all_data.append(data)
        print(f"[{len(all_data)}/{target_count}] {data['用户名']} | {data['星级评分']} | {data['评论内容'][:20]}...")

# ==================== 主程序：自动翻页爬300条 ====================
if __name__ == "__main__":
    start = 0
    while len(all_data) < target_count:
        url = base_url + str(start)
        print(f"\n正在爬取第 {start//10+1} 页：{url}")

        resp = requests.get(url, headers=headers)
        parse_page(resp.text)

        if len(all_data) >= target_count:
            break

        start += 10
        time.sleep(random.uniform(1, 2))

    # 保存 CSV
    df = pd.DataFrame(all_data, columns=csv_columns)
    df.to_csv("豆瓣长影评_300条.csv", index=False, encoding="utf-8-sig")

    # 保存美观 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "影评"
    ws.append(csv_columns)
    for d in all_data:
        ws.append(list(d.values()))

    # 列宽
    col_width = {"A":32, "B":14, "C":10, "D":10, "E":22, "F":85}
    for c, w in col_width.items():
        ws.column_dimensions[c].width = w

    # 自动换行
    align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = align

    wb.save("豆瓣长影评_300条.xlsx")
    print(f"\n✅ 全部完成！共爬取 {len(all_data)} 条长影评")