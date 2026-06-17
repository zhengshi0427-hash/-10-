import requests
import pandas as pd
from lxml import etree
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ==================== 配置（1星影评专用链接） ====================
base_url = "https://movie.douban.com/subject/37116446/reviews?rating=1&start="
target_count = 300  # 采集300条1星影评

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Referer": "https://movie.douban.com/subject/37116446/",
    "Cookie": 'll="118172"; bid=IhGRS2uQwII; dbcl2="295327634:oTNuHdaewsA";'
}

# 6个输出字段
csv_columns = ["用户链接", "用户名", "点赞数", "星级评分", "评论时间", "评论内容"]
all_data = []

star_map = {
    "力荐": "5星",
    "推荐": "4星",
    "还行": "3星",
    "较差": "2星",
    "很差": "1星",
    "未评分": ""
}

# 请求接口获取完整展开影评
def get_full_review(rid):
    try:
        api_url = f"https://movie.douban.com/j/review/{rid}/full"
        resp = requests.get(api_url, headers=headers, timeout=5)
        if resp.status_code == 200:
            html = resp.json()["html"]
            tree = etree.HTML(html)
            return tree.xpath("string()").strip()
    except Exception:
        return ""
    return ""

# ==================== 页面解析函数 ====================
def parse_page(html):
    tree = etree.HTML(html)
    items = tree.xpath('//div[contains(@class,"review-item")]')
    for item in items:
        if len(all_data) >= target_count:
            break
        time.sleep(random.uniform(0.5, 1))
        data = {}
        # 用户主页链接
        data["用户链接"] = item.xpath('.//a[@class="avator"]/@href')[0] if item.xpath('.//a[@class="avator"]/@href') else ""
        # 用户名
        data["用户名"] = item.xpath('.//a[@class="name"]/text()')[0].strip() if item.xpath('.//a[@class="name"]/text()') else ""
        # 有用点赞数
        data["点赞数"] = item.xpath('.//a[contains(@class,"up")]/text()')[0].strip() if item.xpath('.//a[contains(@class,"up")]/text()') else "0"
        # 星级转换
        raw_star = item.xpath('.//span[contains(@class,"allstar")]/@title')
        data["星级评分"] = star_map.get(raw_star[0] if raw_star else "未评分", "")
        # 发布时间
        data["评论时间"] = item.xpath('.//span[@class="main-meta"]/text()')[0].strip() if item.xpath('.//span[@class="main-meta"]/text()') else ""
        # 完整影评内容（接口拉全文）
        rid = item.xpath('.//div[@class="review-short"]/@data-rid')
        if rid:
            data["评论内容"] = get_full_review(rid[0])
        else:
            data["评论内容"] = item.xpath('string(.//div[@class="short-content"])').strip()

        all_data.append(data)
        print(f"[{len(all_data)}/{target_count}] {data['用户名']} | {data['星级评分']} | {data['评论内容'][:22]}...")

# ==================== 自动翻页主逻辑 ====================
if __name__ == "__main__":
    start = 0
    while len(all_data) < target_count:
        req_url = base_url + str(start)
        print(f"\n===== 正在爬取 start={start} 页（1星差评）=====")
        resp = requests.get(req_url, headers=headers)
        parse_page(resp.text)
        # 一页默认10条，偏移+10
        start += 10
        time.sleep(random.uniform(1.2, 2))

    # 导出CSV
    df = pd.DataFrame(all_data, columns=csv_columns)
    df.to_csv("豆瓣_1星差评_300条.csv", index=False, encoding="utf-8-sig")

    # 生成自动换行Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "1星差评影评"
    ws.append(csv_columns)
    for row_data in all_data:
        ws.append(list(row_data.values()))

    # 自定义列宽
    width_cfg = {"A":32, "B":14, "C":10, "D":10, "E":22, "F":88}
    for col, w in width_cfg.items():
        ws.column_dimensions[col].width = w

    # 单元格自动换行
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = wrap_align

    wb.save("豆瓣_1星差评.xlsx")
    print(f"\n✅ 采集完成！实际抓取 {len(all_data)} 条一星影评")